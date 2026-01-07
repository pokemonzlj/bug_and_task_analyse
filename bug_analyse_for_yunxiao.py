# -*- coding: UTF-8 -*-
import tkinter as tk
from tkinter import filedialog
from datetime import datetime
import datetime
# import csv
import openpyxl
import jieba
from operator import itemgetter


class bug_analyse():
    """使用前先将阿里云效缺陷导出为xlsx文件
    字段可以都勾选导出来
    项目中补充了部分自定义字段：
    责任人、测试责任人、缺陷原因/修复方案、解决时间、拒绝时间、系统字段:实际工时"""

    def __init__(self):
        self.bug_info = {'created_bug_count': 0, 'resolved_bug_count': 0, 'online_bug_count': 0, 'offline_bug_count': 0,
                         'rejected_bug_count': 0, 'severity_count': {}, 'resolution_method_count': {}}  # 用于bug维度存储数据
        self.bug_on_person = {}
        # 用户存储个人维度的bug数据，内部结构{'name':{'resolve_bug':10, 'report_bug':0, 'online_response_bug':0,
        # 'offline_response_bug':0, 'reject_bug':0, 'refuse_bug':0, 'offline_test_response_bug',
        # 'online_test_response_bug', 'resolve_time':[1,2,3]}, }
        self.bug_rejected_list = []

    def read_xls_as_dict(self, file_paths=[]):
        result = []
        for file_path in file_paths:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            # 读取首行作为键
            headers = [cell.value for cell in sheet[1]]
            # 逐行读取数据并转换为字典
            for row in sheet.iter_rows(min_row=2, values_only=True):
                entry = dict(zip(headers, row))
                result.append(entry)
        return result

    def select_file(self):
        filepath = filedialog.askopenfilenames()
        return filepath

    def convert_date_format(self, notice="string"):
        """将输入时间转化成标准格式时间"""
        date_string = input(notice)
        current_year = datetime.datetime.now().year
        date_parts = date_string.split('/')
        if len(date_parts) != 2:
            date_parts = date_string.split('-')
        if len(date_parts) != 2:
            date_parts = date_string.split('.')
        if len(date_parts) != 2:
            date_parts = None
        month = int(date_parts[0])
        day = int(date_parts[1])
        if month == 12:
            current_year -= 1
        try:
            converted_date = datetime.datetime(current_year, month, day)
            # print(converted_date.strftime('%Y-%m-%d'))
            # print(type(converted_date))
            return converted_date
        except ValueError:
            return None

    def get_bug_by_date(self, startdate: datetime.datetime, enddate: datetime.datetime, type='create', buglist=[]):
        """通过时间区间筛选bug列表,默认参数必须放在非默认参数之后"""
        need_bug_list = []
        if type == 'create':
            for bug in buglist:
                bug_create_date = bug["创建时间"]
                if not isinstance(bug_create_date, str):
                    bug_create_date = bug_create_date.strftime('%Y-%m-%d %H:%M:%S')
                bug_create_date = bug_create_date.split()[0]
                bug_analyse_date = datetime.datetime.strptime(bug_create_date, '%Y-%m-%d')

                if startdate <= bug_analyse_date <= enddate:
                    need_bug_list.append(bug)
                    self.bug_info['created_bug_count'] += 1
        elif type == 'resolve':
            for bug in buglist:
                # if "解决时间" in bug:
                #     bug_resolved_date = bug["解决时间"]
                #     if bug_resolved_date == ' ':  # 如果没有解决时间，则取一下拒绝时间
                #         bug_rejected_date = bug["拒绝时间"]
                #         if bug_rejected_date != ' ':
                #             bug_rejected_date = bug_rejected_date.split()[0]
                #         if bug_rejected_date != ' ':
                #             bug_analyse_date = datetime.datetime.strptime(bug_rejected_date, '%Y-%m-%d')
                #             if startdate <= bug_analyse_date <= enddate:
                #                 self.bug_info['rejected_bug_count'] += 1
                #                 self.bug_rejected_list.append(bug)
                #         # print(bug_resolved_date)
                #     if bug_resolved_date != ' ':
                #         bug_resolved_date = bug_resolved_date.split()[0]
                #     if bug_resolved_date != ' ':
                #         bug_analyse_date = datetime.datetime.strptime(bug_resolved_date, '%Y-%m-%d')
                #         if startdate <= bug_analyse_date <= enddate:
                #             need_bug_list.append(bug)
                #             self.bug_info['resolved_bug_count'] += 1
                # else:
                    bug_closed_date = bug["完成时间"]
                    bug_reason = bug["缺陷原因/修复方案"]
                    bug_reject_reason = bug["不修复理由"]
                    if bug_closed_date:
                        if isinstance(bug_closed_date, str):
                            bug_analyse_date = datetime.datetime.strptime(bug_closed_date, '%Y-%m-%d')
                        else:
                            bug_analyse_date = bug_closed_date
                        if startdate <= bug_analyse_date <= enddate:
                            # print(bug_reason)
                            if bug_reason:
                                need_bug_list.append(bug)
                                self.bug_info['resolved_bug_count'] += 1
                            elif bug_reject_reason:
                                self.bug_rejected_list.append(bug)
                                self.bug_info['rejected_bug_count'] += 1
        # print(self.bug_rejected_list)
        return need_bug_list

    def bug_level_analyse(self, buglist=[]):
        """bug等级分析"""
        leveldict = {}
        for bug in buglist:
            level = bug['严重程度']
            if level not in leveldict:
                leveldict[level] = 1
            else:
                leveldict[level] += 1
        self.bug_info['severity_count'] = leveldict
        print("BUG按严重程度分类，其中:")
        for level in leveldict:
            print("{}{}个，占比{}；".format(level, leveldict[level], "{:.1f}%".format(
                leveldict[level] / self.bug_info['created_bug_count'] * 100)))
        print("************************************************************")

    def bug_project_analyse(self, buglist=[]):
        """bug归属项目分析"""
        project_dict = {}
        for bug in buglist:
            project = bug['软件平台']
            if project not in project_dict:
                project_dict[project] = 1
            else:
                project_dict[project] += 1
        self.bug_info['project_count'] = project_dict
        print("BUG按项目分类，其中:")
        for project in project_dict:
            print("{}{}个，占比{}；".format(project, project_dict[project], "{:.1f}%".format(
                project_dict[project] / self.bug_info['created_bug_count'] * 100)))
        print("************************************************************")

    def bug_online_analyse(self, buglist=[]):
        """bug线上线下分析"""
        # online = 0
        # offline = 0
        for bug in buglist:
            onoroff = bug['线上/线下']
            if onoroff == '线上':
                self.bug_info['online_bug_count'] += 1
            else:
                self.bug_info['offline_bug_count'] += 1
        print("BUG线上线下分类，其中线上{}个，占比{}；线下{}个，占比{}".format(self.bug_info['online_bug_count'],
                                                                           "{:.1f}%".format(
                                                                               self.bug_info['online_bug_count'] /
                                                                               self.bug_info[
                                                                                   'created_bug_count'] * 100),
                                                                           self.bug_info['offline_bug_count'],
                                                                           "{:.1f}%".format(
                                                                               self.bug_info['offline_bug_count'] /
                                                                               self.bug_info[
                                                                                   'created_bug_count'] * 100)))
        print("************************************************************")

    # def bug_resolution_analyse(self, buglist=[]):
    #     """bug解决方法分析"""
    #     resolution = {}
    #     resolve_num = 0
    #     for bug in buglist:
    #         bugstatus = bug['状态']
    #         if '已关闭' in bugstatus or '已修复' in bugstatus or '暂不修复' in bugstatus:
    #             res = bug['解决方法']
    #             if res not in resolution:
    #                 resolution[res] = 1
    #             else:
    #                 resolution[res] += 1
    #             resolve_num += 1
    #     self.bug_info['resolution_method_count'] = resolution
    #     print("BUG按解决方法分类，其中:")
    #     for res in resolution:
    #         print("{}{}个，占比{}；".format(res, resolution[res], "{:.1f}%".format(
    #             resolution[res] / resolve_num * 100)))
    #     print("************************************************************")

    def bug_reject_analyse(self, buglist=[]):
        """bug拒绝人分析"""
        for bug in buglist:
            transactor = bug['负责人']
            if transactor not in self.bug_on_person:
                self.bug_on_person[transactor] = {}
                self.bug_on_person[transactor]['reject_bug'] = 1
            else:
                if 'reject_bug' not in self.bug_on_person[transactor]:
                    self.bug_on_person[transactor]['reject_bug'] = 0
                self.bug_on_person[transactor]['reject_bug'] += 1
        print("BUG按拒绝人统计，其中:")
        for res in self.bug_on_person:
            if 'reject_bug' in self.bug_on_person[res]:
                print("{}拒绝{}个，占比{}；".format(res, self.bug_on_person[res]['reject_bug'],
                                                  "{:.1f}%".format(
                                                      self.bug_on_person[res]['reject_bug'] / self.bug_info[
                                                          'rejected_bug_count'] * 100)))
        print("************************************************************")

    def bug_report_analyse(self, buglist=[]):
        """bug报告人分析"""
        for bug in buglist:
            reporter = bug['创建者']
            if reporter not in self.bug_on_person:
                self.bug_on_person[reporter] = {}
                self.bug_on_person[reporter]['report_bug'] = 1
            else:
                self.bug_on_person[reporter]['report_bug'] += 1
        print("BUG按报告人统计，其中:")
        for res in self.bug_on_person:
            if 'report_bug' in self.bug_on_person[res]:
                print("{}报告{}个，占比{}；".format(res, self.bug_on_person[res]['report_bug'],
                                                  "{:.1f}%".format(
                                                      self.bug_on_person[res]['report_bug'] / self.bug_info[
                                                          'created_bug_count'] * 100)))
        print("************************************************************")

    def bug_responsible_analyse(self, buglist=[]):
        """关闭状态bug测试责任人、责任人和解决人分析"""
        # responsible = {}
        # global test_responser
        print("分析{}条缺陷记录".format(len(buglist)))
        for bug in buglist:
            level = bug['严重程度']
            onoroff = bug['线上/线下']
            solver = bug['负责人']
            # if '责任人' in bug:  #如果责任人存在就取责任人
            bugstatus = bug['状态']
            responser = bug['责任人']
            test_responser = bug['测试责任人']
            # if test_responser:  # 如果测试责任人不为空
            #     test_responser = test_responser.replace(";", '')
            # if responser:  # 如果责任人不为空
            #     responser = responser.replace(";", '')
            if responser == ' ' and solver != ' ':  # 责任人为空则取解决人为责任人
                responser = solver
            if onoroff == "线上":
                # if level != '4-轻微':
                if bugstatus == '已关闭':  # or bugstatus == '已修复'
                    if solver != ' ':
                        if solver not in self.bug_on_person:
                            self.bug_on_person[solver] = {}
                            self.bug_on_person[solver]['resolve_bug'] = 1
                        else:
                            if 'resolve_bug' not in self.bug_on_person[solver]:
                                self.bug_on_person[solver]['resolve_bug'] = 0
                            self.bug_on_person[solver]['resolve_bug'] += 1
                # else:
                #     responser = bug['负责人']
                    if responser:  # 区分缺陷等级保存
                        if responser not in self.bug_on_person:
                            self.bug_on_person[responser] = {}
                            self.bug_on_person[responser]['online_response_bug'] = 1
                        else:
                            if 'online_response_bug' not in self.bug_on_person[responser]:
                                self.bug_on_person[responser]['online_response_bug'] = 0
                            self.bug_on_person[responser]['online_response_bug'] += 1
                        if 'online_bug' not in self.bug_on_person[responser]:
                            self.bug_on_person[responser]['online_bug'] = {}
                        if level not in self.bug_on_person[responser]['online_bug']:
                            self.bug_on_person[responser]['online_bug'][level] = 0
                        self.bug_on_person[responser]['online_bug'][level] += 1
                    if test_responser:
                        if test_responser not in self.bug_on_person:
                            self.bug_on_person[test_responser] = {}
                            self.bug_on_person[test_responser]['online_test_response_bug'] = 1
                        else:
                            if 'online_test_response_bug' not in self.bug_on_person[test_responser]:
                                self.bug_on_person[test_responser]['online_test_response_bug'] = 0
                            self.bug_on_person[test_responser]['online_test_response_bug'] += 1
                        if 'online_bug_response' not in self.bug_on_person[test_responser]:
                            self.bug_on_person[test_responser]['online_bug_response'] = {}
                        if level not in self.bug_on_person[test_responser]['online_bug_response']:
                            self.bug_on_person[test_responser]['online_bug_response'][level] = 0
                        self.bug_on_person[test_responser]['online_bug_response'][level] += 1
            elif bugstatus == '已关闭':   # 如果是线下已关闭bug，都算
                if solver != ' ':
                    if solver not in self.bug_on_person:
                        self.bug_on_person[solver] = {}
                        self.bug_on_person[solver]['resolve_bug'] = 1
                    else:
                        if 'resolve_bug' not in self.bug_on_person[solver]:
                            self.bug_on_person[solver]['resolve_bug'] = 0
                        self.bug_on_person[solver]['resolve_bug'] += 1
                if responser:
                    # if responser not in self.bug_on_person:
                    #     self.bug_on_person[responser] = {}
                    #     self.bug_on_person[responser]['offline_response_bug'] = 1
                    # else:
                    #     if 'offline_response_bug' not in self.bug_on_person[responser]:
                    #         self.bug_on_person[responser]['offline_response_bug'] = 0
                    #     self.bug_on_person[responser]['offline_response_bug'] += 1
                    if responser not in self.bug_on_person:
                        self.bug_on_person[responser] = {}
                        self.bug_on_person[responser]['offline_response_bug'] = 1
                    else:
                        if 'offline_response_bug' not in self.bug_on_person[responser]:
                            self.bug_on_person[responser]['offline_response_bug'] = 0
                        self.bug_on_person[responser]['offline_response_bug'] += 1
                    if 'offline_bug' not in self.bug_on_person[responser]:
                        self.bug_on_person[responser]['offline_bug'] = {}
                    if level not in self.bug_on_person[responser]['offline_bug']:
                        self.bug_on_person[responser]['offline_bug'][level] = 0
                    self.bug_on_person[responser]['offline_bug'][level] += 1
                if test_responser:
                    if test_responser not in self.bug_on_person:
                        self.bug_on_person[test_responser] = {}
                        self.bug_on_person[test_responser]['offline_test_response_bug'] = 1
                    else:
                        if 'offline_test_response_bug' not in self.bug_on_person[test_responser]:
                            self.bug_on_person[test_responser]['offline_test_response_bug'] = 0
                        self.bug_on_person[test_responser]['offline_test_response_bug'] += 1
        print("BUG按解决人统计，其中:")
        for res in self.bug_on_person:
            if 'resolve_bug' in self.bug_on_person[res]:
                print("{}解决{}个；".format(res, self.bug_on_person[res]['resolve_bug']))
        print("************************************************************")
        print("BUG按责任人统计，其中:")
        for res in self.bug_on_person:
            if 'offline_response_bug' in self.bug_on_person[res]:
                if 'online_response_bug' in self.bug_on_person[res]:
                    print("{}责任归属线上{}个，线下{}个；".format(res, self.bug_on_person[res]['online_response_bug'],
                                                          self.bug_on_person[res]['offline_response_bug']))
                    for level in self.bug_on_person[res]['online_bug']:
                        print("其中线上{}:{}个".format(level, self.bug_on_person[res]['online_bug'][level]))
                    for level in self.bug_on_person[res]['offline_bug']:
                        print("其中线下{}:{}个".format(level, self.bug_on_person[res]['offline_bug'][level]))
                else:
                    print("{}责任归属线上0个，线下{}个；".format(res, self.bug_on_person[res]['offline_response_bug']))
                    for level in self.bug_on_person[res]['offline_bug']:
                        print("其中线下{}:{}个".format(level, self.bug_on_person[res]['offline_bug'][level]))
        print("************************************************************")
        print("BUG按测试责任人统计，其中:")
        for res in self.bug_on_person:
            if 'online_test_response_bug' in self.bug_on_person[res]:
                print("{}线上bug测试责任{}个；".format(res, self.bug_on_person[res]['online_test_response_bug']))
                for level in self.bug_on_person[res]['online_bug_response']:
                    print("其中线上{}:{}个".format(level, self.bug_on_person[res]['online_bug_response'][level]))
        print("************************************************************")

    def bug_deal_time_analyse(self, buglist=[]):
        """bug花费时间的统计"""
        for bug in buglist:
            resolver = bug['负责人']
            resolve_time = bug['实际工时汇总']
            responser = bug['责任人']
            if resolve_time:
                resolve_time = float("{:.1f}".format(float(resolve_time)))
                if responser:  # 如果责任人不为空
                    responser = responser.replace(";", '')
                if resolver not in self.bug_on_person:
                    self.bug_on_person[resolver] = {}
                    if 'resolve_total_time' not in self.bug_on_person[resolver]:
                        self.bug_on_person[resolver]['resolve_total_time'] = 0
                    self.bug_on_person[resolver]['resolve_total_time'] += resolve_time
                else:
                    if 'resolve_total_time' not in self.bug_on_person[resolver]:
                        self.bug_on_person[resolver]['resolve_total_time'] = 0
                    self.bug_on_person[resolver]['resolve_total_time'] += resolve_time
                if resolver != responser:
                    if 'help_resolve_time' not in self.bug_on_person[resolver]:
                        self.bug_on_person[resolver]['help_resolve_time'] = 0
                    self.bug_on_person[resolver]['help_resolve_time'] += resolve_time
        print("BUG按花费时间统计，其中:")
        for res in self.bug_on_person:
            if 'resolve_total_time' in self.bug_on_person[res]:
                if "help_resolve_time" in self.bug_on_person[res]:
                    print("{}修复bug总时间{:.1f}小时,其中帮助他人修复bug总时间{:.1f}小时；".format(res, self.bug_on_person[res]
                    ['resolve_total_time'],self.bug_on_person[res]['help_resolve_time']
                                                      ))
                else:
                    print("{}修复bug总时间{:.1f}小时,其中帮助他人修复bug总时间{:.1f}小时；".format(res, self.bug_on_person[res][
                        'resolve_total_time'], 0))
        print("************************************************************")

    def bug_resolve_time_analyse(self, buglist=[]):
        """线上bug处理时间和处理方式分析"""
        for bug in buglist:
            report_time = bug['创建时间']
            resolve_time = bug['解决时间']
            onoroff = bug['线上/线下']
            if onoroff == "线上":
                if not resolve_time or resolve_time == ' ':
                    continue
                time_format = "%Y-%m-%d %H:%M:%S"
                if not isinstance(report_time, str):
                    report_time = report_time.strftime(time_format)
                # time_start = report_time.split()[0]
                time_start = report_time
                time_start = datetime.datetime.strptime(time_start, '%Y-%m-%d %H:%M:%S')
                if not isinstance(resolve_time, str):
                    resolve_time = resolve_time.strftime(time_format)
                # time_end = resolve_time.split()[0]
                time_end = resolve_time
                time_end = datetime.datetime.strptime(time_end, '%Y-%m-%d %H:%M:%S')
                # 计算两个时间之间的差值
                time_difference = time_end - time_start
                # 获取以天为单位的差值
                days_difference = time_difference.days
                solver = bug['负责人']
                if solver not in self.bug_on_person:
                    self.bug_on_person[solver] = {}
                # resolve_way = bug['解决方法']
                # if '拒绝' in resolve_way:
                # if 'refuse_time' not in self.bug_on_person[solver]:
                #     self.bug_on_person[solver]['refuse_time'] = 0
                # self.bug_on_person[solver]['refuse_time'] += 1
                if 'resolve_time' not in self.bug_on_person[solver]:
                    self.bug_on_person[solver]['resolve_time'] = []
                self.bug_on_person[solver]['resolve_time'].append(days_difference)
                verifier = bug['验证者']
                close_time = bug['完成时间']
                if not isinstance(close_time, str):
                    close_time = close_time.strftime(time_format)
                close_time = datetime.datetime.strptime(close_time, '%Y-%m-%d %H:%M:%S')
                # time_end_day = resolve_time.split()[0]
                # time_end_day = datetime.datetime.strptime(time_end_day, '%Y-%m-%d')
                verification_time_difference = close_time - time_end
                # 获取以天为单位的差值
                verification_days_difference = verification_time_difference.days
                if verifier not in self.bug_on_person:
                    self.bug_on_person[verifier] = {}
                if 'verification_time' not in self.bug_on_person[verifier]:
                    self.bug_on_person[verifier]['verification_time'] = []
                self.bug_on_person[verifier]['verification_time'].append(verification_days_difference)
        print("BUG按修复时间统计，其中:")
        for res in self.bug_on_person:
            if 'resolve_time' in self.bug_on_person[res]:
                print("{}平均修复bug时间{}天；最大缺陷响应天数{}天；".format(res, "{:.1f}".format(
                    sum(self.bug_on_person[res]['resolve_time']) /
                    len(self.bug_on_person[res]['resolve_time'])), max(self.bug_on_person[res]['resolve_time'])))
        # print("************************************************************")
        # print("BUG按拒绝次数统计，其中:")
        # for res in self.bug_on_person:
        #     if 'refuse_time' in self.bug_on_person[res]:
        #         print("{}共拒绝bug{}个；".format(res, self.bug_on_person[res]['refuse_time']))
        print("************************************************************")
        for res in self.bug_on_person:
            if 'verification_time' in self.bug_on_person[res]:
                print("{}最大已修复缺陷响应时间{}天；".format(res, max(self.bug_on_person[res]['verification_time'])))
        print("************************************************************")

    def bug_reopen_analyse(self, buglist=[]):
        """bug反复激活分析"""
        print("BUG多次激活列表:")
        print("Bug编号 Bug标题 责任人 严重程度 激活次数")
        for bug in buglist:
            bugreopentime = bug['激活次数']
            if bugreopentime == 0 or bugreopentime == '0':
                pass
            else:
                bugid = bug['Bug编号']
                bugtitle = bug['Bug标题']
                buglevel = bug['严重程度']
                bugstatus = bug['状态']
                if '已关闭' in bugstatus or '解决' in bugstatus:
                    bugresponsible = bug['负责人']
                else:
                    bugresponsible = bug['负责人']
                print("{} {} {} {} {}".format(bugid, bugtitle, bugresponsible, buglevel, bugreopentime))

                # 记录个人激活次数
                if bugresponsible not in self.bug_on_person:
                    self.bug_on_person[bugresponsible] = {}
                if 'reopen_count' not in self.bug_on_person[bugresponsible]:
                    self.bug_on_person[bugresponsible]['reopen_count'] = 0
                self.bug_on_person[bugresponsible]['reopen_count'] += int(bugreopentime)
        print("************************************************************")

    def bug_reason_analyse(self, buglist=[]):
        """缺陷产生原因拆词分析"""
        words_freq = {}
        for bug in buglist:
            onoroff = bug['线上/线下']
            reason = bug['缺陷原因/修复方案']
            if not reason:
                continue
            if onoroff == "线上":
                print("'" + reason + "'")
            words = jieba.lcut(reason, cut_all=True)  # 全模式分词
            # words = jieba.lcut_for_search(reason)  # 适用于搜索引擎的分词模式，会对长词再次切分
            for word in words:
                if word in words_freq:
                    words_freq[word] += 1
                else:
                    words_freq[word] = 1
        # 按词频大小排序词频字典，并返回一个按词频降序排列的列表
        sorted_word_freq = sorted(words_freq.items(), key=itemgetter(1), reverse=True)
        for word, freq in sorted_word_freq:
            print(f"{word}: {freq} 次")

    def judge_weekday(self, year=2024, month=10, day=8):
        """判断日期是星期几"""
        date = datetime.datetime(year, month, day)
        weekday_num = date.weekday() + 1
        print("日期对应的是星期{}".format(weekday_num))

    def total_analyse(self):
        filepath = self.select_file()
        buglist = self.read_xls_as_dict(filepath)
        startdate = self.convert_date_format("请输入开始时间，格式可以是月/日或月-日或月.日：")
        enddate = self.convert_date_format("请输入结束时间，格式可以是月/日或月-日或月.日：")
        create_buglist = self.get_bug_by_date(startdate, enddate, 'create', buglist)
        resolve_buglist = self.get_bug_by_date(startdate, enddate, 'resolve', buglist)
        # self.total_bug_count = len(buglist)
        print("{}至{}区间内共新建{}个bug，修复{}个bug，拒绝{}个bug".format(startdate.strftime('%Y-%m-%d'),
                                                                         enddate.strftime('%Y-%m-%d'),
                                                                         self.bug_info['created_bug_count'],
                                                                         self.bug_info['resolved_bug_count']
                                                                         , self.bug_info['rejected_bug_count']))
        self.bug_level_analyse(create_buglist)
        self.bug_online_analyse(create_buglist)
        self.bug_project_analyse(create_buglist)
        # self.bug_resolution_analyse(resolve_buglist)
        self.bug_report_analyse(create_buglist)
        self.bug_reject_analyse(self.bug_rejected_list)
        # print("缺陷创建列表中")
        # self.bug_responsible_analyse(create_buglist)
        print("缺陷修复列表中")
        self.bug_responsible_analyse(resolve_buglist)
        self.bug_resolve_time_analyse(resolve_buglist)
        self.bug_deal_time_analyse(resolve_buglist)
        # self.bug_reopen_analyse(resolve_buglist)
        self.bug_reason_analyse(resolve_buglist)
        self.export_person_data()

    def export_person_data(self):
        """将人员统计数据导出到Excel"""
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "人员统计"

        headers = ["姓名", "报告Bug数", "解决Bug数", "拒绝Bug数", "线上责任Bug数",
                   "线上-致命", "线上-严重", "线上-一般", "线上-轻微",
                   "线下责任Bug数",
                   "线下-致命", "线下-严重", "线下-一般", "线下-轻微",
                   "线上测试责任Bug数",
                   "本次测试-致命", "本次测试-严重", "本次测试-一般", "本次测试-轻微",
                   "平均修复时间(天)",
                   "修复总工时(小时)", "帮助修复工时(小时)", "激活次数"]
        sheet.append(headers)

        # 统一使用用户指定的档位：致命, 严重, 一般, 轻微
        target_levels = ["致命", "严重", "一般", "轻微"]
        
        # 兼容性映射（以防原始数据包含 "1-致命" 等格式）
        # 这里在统计时直接进行映射

        for name, data in self.bug_on_person.items():
            if not name or name == ' ': continue
            report_bug = data.get('report_bug', 0)
            resolve_bug = data.get('resolve_bug', 0)
            
            # 责任人统计 (兼容性处理)
            def get_mapped_data(d):
                mapped = {lvl: 0 for lvl in target_levels}
                for k, v in d.items():
                    if "致命" in k: mapped["致命"] += v
                    elif "严重" in k: mapped["严重"] += v
                    elif "一般" in k: mapped["一般"] += v
                    elif "轻微" in k: mapped["轻微"] += v
                return mapped

            online_mapped = get_mapped_data(data.get('online_bug', {}))
            offline_mapped = get_mapped_data(data.get('offline_bug', {}))
            
            online_total = sum(online_mapped.values())
            online_split = [online_mapped[lvl] for lvl in target_levels]
            
            offline_total = sum(offline_mapped.values())
            offline_split = [offline_mapped[lvl] for lvl in target_levels]

            # 线上 测试责任人 统计
            online_test_mapped = get_mapped_data(data.get('online_bug_response', {}))
            online_test_total = sum(online_test_mapped.values())
            online_test_split = [online_test_mapped[lvl] for lvl in target_levels]
            
            reject_bug = data.get('reject_bug', 0)

            # 时间统计
            avg_resolve_time = 0
            if 'resolve_time' in data and data['resolve_time']:
                avg_resolve_time = sum(data['resolve_time']) / len(data['resolve_time'])

            total_work_time = data.get('resolve_total_time', 0)
            help_work_time = data.get('help_resolve_time', 0)
            reopen_count = data.get('reopen_count', 0)

            row = [name, report_bug, resolve_bug, reject_bug, online_total] + online_split + \
                  [offline_total] + offline_split + \
                  [online_test_total] + online_test_split + \
                  [round(avg_resolve_time, 2),
                   total_work_time, help_work_time, reopen_count]
            sheet.append(row)

        filename = "person_analyse_yunxiao.xlsx"
        wb.save(filename)
        print(f"人员分析数据已导出至 {filename}")


if __name__ == "__main__":
    total_analyse = bug_analyse()
    total_analyse.total_analyse()
